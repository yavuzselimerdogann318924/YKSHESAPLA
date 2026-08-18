#!/usr/bin/env python3
"""Convert ÖSYM Table 3/4 workbooks into a compact browser dataset."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCORE_TYPE_CODES = {"TYT": 0, "SAY": 1, "EA": 2, "SÖZ": 3, "DİL": 4}
UNIVERSITY_TYPE_CODES = {
    "DEVLET": 0,
    "VAKIF": 1,
    "KKTC": 2,
    "YURTDISI KAMU": 3,
    "YURTDISI VAKIF": 4,
}


def number(value: Any) -> int | float | None:
    if not isinstance(value, (int, float)):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


class Dictionary:
    def __init__(self) -> None:
        self.values: list[str] = []
        self.indices: dict[str, int] = {}

    def add(self, value: Any) -> int:
        text = str(value or "").strip()
        if text not in self.indices:
            self.indices[text] = len(self.values)
            self.values.append(text)
        return self.indices[text]


def import_workbook(
    path: Path,
    level: int,
    universities: Dictionary,
    faculties: Dictionary,
    programs: Dictionary,
) -> list[list[Any]]:
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    imported: list[list[Any]] = []

    for source in worksheet.iter_rows(min_row=4, values_only=True):
        code = str(source[0] or "").strip()
        if not code.isdigit():
            continue

        score_type = str(source[5] or "").strip()
        university_type = str(source[1] or "").strip()
        row: list[Any] = [
            code,
            level,
            UNIVERSITY_TYPE_CODES[university_type],
            universities.add(source[2]),
            faculties.add(source[3]),
            programs.add(source[4]),
            SCORE_TYPE_CODES[score_type],
        ]

        for start in (6, 10, 14, 18):
            row.extend(number(value) for value in source[start : start + 4])

        imported.append(row)

    return imported


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--associate", required=True, type=Path)
    parser.add_argument("--bachelor", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    universities = Dictionary()
    faculties = Dictionary()
    programs = Dictionary()

    associate_rows = import_workbook(
        args.associate, 2, universities, faculties, programs
    )
    bachelor_rows = import_workbook(
        args.bachelor, 4, universities, faculties, programs
    )

    payload = {
        "version": 1,
        "year": 2026,
        "universities": universities.values,
        "faculties": faculties.values,
        "programs": programs.values,
        "rows": associate_rows + bachelor_rows,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "rows": len(payload["rows"]),
                "universities": len(universities.values),
                "faculties": len(faculties.values),
                "programs": len(programs.values),
                "bytes": args.output.stat().st_size,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
